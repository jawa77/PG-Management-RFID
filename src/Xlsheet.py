
from openpyxl import Workbook, load_workbook 

class Xlsheet:
    @staticmethod
    def writexl(name,date,time):

        # Load an existing workbook or create a new one if it doesn't exist
        try:
            workbook = load_workbook("attendance.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "Name"
            sheet["B1"] = "Date"
            sheet["C1"] = "Time"

        # Check if the user and date combination already exists in the sheet
        row_idx = 2
        while sheet[f"A{row_idx}"].value is not None:
            if sheet[f"A{row_idx}"].value == name and sheet[f"B{row_idx}"].value == date:
                # Append the time to the existing entry
                existing_time = sheet[f"C{row_idx}"].value
                sheet[f"C{row_idx}"] = f"{existing_time} {time}"
                break
            row_idx += 1
        else:
            # User and date combination not found, add a new row
            sheet[f"A{row_idx}"] = name
            sheet[f"B{row_idx}"] = date
            sheet[f"C{row_idx}"] = time

        # Save the workbook with the new data
        workbook.save("attendance.xlsx")

